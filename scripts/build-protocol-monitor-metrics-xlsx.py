#!/usr/bin/env python3
"""Generate an Excel workbook documenting every metric shown in the
Protocol Monitor (Usage Overview + 6 direction tabs + Requests/Responses).

Chinese text mirrors the in-app explainer popups (condensed).
"""

from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


COLUMNS = ["Metric", "Meaning", "How Computed", "Typical Values", "Additional Notes"]

ROWS: list[tuple[str, str, str, str, str]] = [
    # -------------------------------------------------------------------
    # Usage Overview
    # -------------------------------------------------------------------
    (
        "UsageOverview.Messages",
        "当前 usage 时间窗口内 user + assistant message 的总数,按 role 分类计数。",
        "从磁盘上每个 session 的 transcript 文件读取,每一条 message.role === \"user\" 计 1 次,每一条 assistant 记录 1 次;Total = User + Assistant。",
        "user ≈ assistant 的单轮 session:Total = 2 × 轮数;多 tool-call 的 agent 任务 assistant 常大于 user。",
        "system prompt、tool role 的 message 不计入。聚合发生在 gateway 方法 sessions.usage → scanTranscriptFile。启用 model 过滤后,会只基于匹配该 model 的 session 重新聚合。",
    ),
    (
        "UsageOverview.Throughput",
        "session 活跃期间 token 流动的速率(tokens/min),只算 session 实际在工作的那段时间,不是自然时间。",
        "throughput = totalTokens / (durationSumMs / 60000);durationSumMs 来自每个 session 上报的 usage.durationMs(transcript 活动跨越的时间)之和。",
        "一般云端 LLM:1k–20k tokens/min。纯 cache read 或 cache hit 高的 agent 可能显示非常高的值。",
        "没有可测量时长的 session 被跳过(没法算速率)。session 内部空闲间隙仍计入其时长。Token 包括 input/output/cacheRead/cacheWrite,与 Total Tokens 同源。",
    ),
    (
        "UsageOverview.ToolCalls",
        "assistant 调用 tool 的总次数。一次 assistant 回复里调用了 3 个 tool,对这个数字贡献 3。",
        "scanTranscriptFile 对每条 assistant 记录提取 entry.toolNames,其长度加到 messageCounts.toolCalls;同时 per-tool map 各自计数。",
        "1 轮任务常见几十到几百次 tool call;探索任务可能数千。",
        "Unique tools = per-tool map 的 size。Tool Results 与 Tool Calls 正常 1:1,但流式取消、出错或 aborted 的回合会让两个数字对不上。",
    ),
    (
        "UsageOverview.AvgTokens",
        "时间窗口内每一条 user + assistant message 平均搬运了多少 token。",
        "avgTokens = round(totalTokens / messages.total);totalTokens = input + output + cacheRead + cacheWrite。",
        "轻量对话 ~1k / msg;带大量缓存 transcript 的 agent 数万 / msg 不罕见。",
        "大量 cache read 会让总 token 快速上涨,平均值也会很高。messages.total = 0 时卡片显示 0(不除零)。",
    ),
    (
        "UsageOverview.CacheHit",
        "prompt 侧的 token 中,有多少比例从 model 的 prompt cache 命中,而不是作为全新 input 重新计费。",
        "cacheHitRate = cacheRead / (input + cacheRead),百分比表示。分母只含 prompt 侧,不含 output 和 cacheWrite。",
        "Good ≥ 60%(绿);30–60% 一般(黄);< 30% 基本没命中(红)。",
        "与 cache write 是两个量:cacheWrite 是为后续轮次写入缓存的 token。",
    ),
    (
        "UsageOverview.ErrorRate",
        "带 error 信号的 message 数相对于 message 总数的比例。",
        "errorRate = errors / messages.total,百分比。errors 来自 scanTranscriptFile:① tool result 被标记为 error +1;② assistant 轮次 stopReason ∈ {error, aborted, timeout} +1。",
        "Good ≤ 1%;1–5% 警告;> 5% 严重。",
        "一轮可能同时记录多个 tool result error + error stopReason,极端情况下比率可能超过 100%。",
    ),
    (
        "UsageOverview.TotalCost",
        "时间窗口内每条 model usage 记录报告的 per-session 费用总和(USD)。",
        "每条 assistant usage 的 cost 拆分(input/output/cacheRead/cacheWrite)由 applyCostBreakdown 累加到 session.totalCost;sessions.usage 再把所有匹配 session 的 totalCost 相加。",
        "单次短 agent task 几美分;长流程几到数十美元。",
        "副标题 Per message = totalCost / messages.total(保留 4 位小数)。本地运行或价格未知的 usage 不计入 cost(会出现 missingCost 条目),但 token 总量仍会计入。",
    ),
    (
        "UsageOverview.TotalTokens",
        "当前时间窗口内归属的全部 token,按四类累加。",
        "totalTokens = input + output + cacheRead + cacheWrite。每个 session 若上报了 total 则直接用,否则四项相加;之后 gateway 跨 session 累加。",
        "单轮短对话 < 10k;长 agent run 1–10M 都有。",
        "Input:重新 tokenize 的 prompt;Output:assistant 生成;Cache read:命中缓存的 prompt;Cache write:写入缓存的 prompt。Avg Tokens = 该 total ÷ message 数;Cache Hit 只用 input+cacheRead。",
    ),
]


def throughput_rows(direction_prefix: str, short_label: str, direction_zh: str, transport_zh: str) -> list[tuple[str, str, str, str, str]]:
    """Generate the 4 throughput cards for one direction tab."""
    return [
        (
            f"{direction_prefix}.PeakThroughput",
            f"{short_label} 方向上,任意一个 2 秒采样桶里瞬时字节速率的最大值 —— 这一段时间最 “尖” 的那一下流量。",
            "每条 trace record 带 ts 与 payloadSize;computeNetworkStats 按 floor(ts / 2000) * 2000 分到 2 秒 bucket 累加字节,再用 bytes / 2000 * 1000 换算成 bytes/s,peak = max(bucket.bytesPerSec)。",
            "空闲 < 1 KB/s;轻量 RPC 几 KB/s;LLM 请求体 burst 可到 MB/s。",
            f"Transport:{transport_zh}。Peak 只反映最高那一桶;稳态压力看 Average。",
        ),
        (
            f"{direction_prefix}.AverageThroughput",
            f"{short_label} 方向上所有采样桶 bytes/s 的算术平均。",
            "average = sum(bucket.bytesPerSec) / bucket count。分母是 bucket 数,不是总秒数 —— 完全没有流量的时段没有 bucket,不会拉低均值。",
            "链路大多数时候是否闲置可对比 Peak 判断:peak ≫ avg 说明抖动大。",
            f"Transport:{transport_zh}。要算“自 session 开始的真实平均”更准确的做法是 totalBytes / 实际持续时长。",
        ),
        (
            f"{direction_prefix}.TotalBytes",
            f"{short_label} 方向上,自 UI 进程启动以来所有 trace record 的 payload 字节累加和。",
            "controller 侧持久 accumulator(按 trace id 去重)累加每条匹配 trace 的 payloadSize,与 ring buffer eviction 解耦,不会随旧 trace 被挤出而回落。",
            "轻量 RPC 几 MB;LLM 任务 agent↔model 方向累计到几十上百 MB。",
            f"Transport:{transport_zh}。这是 UI 自进程启动的累计,不是 gateway 端计费;重置 Protocol Monitor 后归零。",
        ),
        (
            f"{direction_prefix}.LastActivity",
            f"{short_label} 方向上最后一个有数据的 2 秒 bucket 的开始时间,以及那一桶里的瞬时速率。",
            "取 samples 中 ts 最大的那一桶,显示 floor(ts / 2000) * 2000 对应的本地时间、bytesPerSec 和 bucket 内累计 bytes。",
            "活跃链路应在最近几秒内;空闲链路可能是几分钟前。",
            "当 accumulator 里没有该方向的任何 bucket 时显示 “—”。",
        ),
    ]


def latency_rows(direction_prefix: str, latency_name: str, latency_intro_zh: str, how_measured_zh: str) -> list[tuple[str, str, str, str, str]]:
    return [
        (
            f"{direction_prefix}.{latency_name}.Avg",
            f"{latency_name} 所有采样的算术平均。{latency_intro_zh}",
            "avg = sum(latencyMs) / count。",
            "LAN WS 一般 < 10 ms;跨 Tailscale 几十 ms;公网几十到几百 ms。",
            how_measured_zh + " 少量极慢请求(冷启动、超时重试)会拉高 avg;看 p50 更能反映常态。",
        ),
        (
            f"{direction_prefix}.{latency_name}.p50",
            f"{latency_name} 的中位数。把所有 sample 按 latency 升序排,中间那个点的值。{latency_intro_zh}",
            "sorted[floor(count * 0.5)]。",
            "稳定链路下应接近 avg;长尾场景下 p50 会明显小于 avg。",
            how_measured_zh + " p50 不会被极端值拉偏;p50 ≪ avg → 分布有长尾。",
        ),
        (
            f"{direction_prefix}.{latency_name}.p95",
            f"{latency_name} 的 95% 百分位数;100 个请求里 95 个比它快、5 个比它慢。{latency_intro_zh}",
            "sorted[floor(count * 0.95)]。",
            "性能优化通常以压低 p95 为目标,比如 LAN RPC 要 p95 < 50 ms。",
            how_measured_zh + " sample < 20 时抖动大,主要看趋势而非绝对值。",
        ),
        (
            f"{direction_prefix}.{latency_name}.Peak",
            f"{latency_name} 中最慢的那一笔 sample。{latency_intro_zh}",
            "sorted[count - 1] = max(latencyMs)。",
            "通常是一次冷启动 / 网络抖动 / 极大 prompt 的代价,几秒到几十秒都可能。",
            how_measured_zh + " 当作“知道一下最坏到过多少”的参考,不要拿来当 SLA。",
        ),
    ]


TRANSPORT_OP_GW = "Operator(浏览器或 Mac/Windows 客户端)↔ Gateway,WebSocket 长连接。Throughput 统计的是 WS 帧的 payload 字节。"
TRANSPORT_GW_NODE = "Gateway ↔ Node(本机或远端 PC),WebSocket 长连接,带 OpenClaw 自己的 RPC / event 协议。Throughput 统计的是 WS 帧的 payload 字节。"
TRANSPORT_AGENT_MODEL = "Agent → Model 走 HTTP/HTTPS(不是 WebSocket)。每次 LLM 调用 = 一次 HTTPS POST(system prompt、transcript、tool schema 等在请求体里)。Throughput 记录的是这次 HTTP 请求体的 payload 字节,由 agent 在发出请求时打点 trace。"
TRANSPORT_MODEL_AGENT = "Model → Agent 是 SSE(Server-Sent Events)流,承载在 Agent→Model 那次 HTTPS 响应之上。Model 把 assistant token、tool call、stop_reason 以一连串 SSE 事件回流。Throughput 记录的是每个 SSE 事件帧的 payload 字节。"


# ---------------------------------------------------------------------------
# Operator ↔ Gateway
# ---------------------------------------------------------------------------
ROWS.extend(throughput_rows("Operator->Gateway", "Operator → Gateway", "op→gw", TRANSPORT_OP_GW))
ROWS.extend(
    latency_rows(
        "Operator->Gateway",
        "OneWayLatency",
        "这里测量的是 Operator → Gateway 单向 wire latency:operator 把 WS frame 发出的瞬间(frame.sentAt)到 gateway 收到的瞬间(Date.now())之间的差值。",
        "打点在 protocol-trace-store.captureTrace:recvTs − sentAt,存到 trace.oneWayLatencyMs,UI 端 computeWsOneWayLatency 按 source 聚合。前提是时钟 NTP 同步(gateway 会把负值钳到 0、> 60s 的丢弃)。",
    )
)

ROWS.extend(throughput_rows("Gateway->Operator", "Gateway → Operator", "gw→op", TRANSPORT_OP_GW))
ROWS.extend(
    latency_rows(
        "Gateway->Operator",
        "OneWayLatency",
        "这里显示的是 Gateway → Operator 方向的 latency。由于 gateway 拿不到 operator 端的 recv 时间,对称地借用反向(op→gw)的 one-way 数据估计;LAN / Tailscale 上通常上下行延迟接近。",
        "打点公式 recvTs − sentAt 只能在接收端跑,outbound 无法直接测量。目前展示反向 one-way 数据作为近似;要精确测量需要 peer 周期性回报自己观测到的 gateway→peer one-way。",
    )
)

# ---------------------------------------------------------------------------
# Gateway ↔ Node
# ---------------------------------------------------------------------------
ROWS.extend(throughput_rows("Gateway->Node", "Gateway → Node", "gw→node", TRANSPORT_GW_NODE))
ROWS.extend(
    latency_rows(
        "Gateway->Node",
        "OneWayLatency",
        "这里显示的是 Gateway → Node 方向的 latency。由于 gateway 拿不到 node 端的 recv 时间,对称地借用反向(node→gw)的 one-way 数据估计。",
        "打点公式 recvTs − sentAt 只能在接收端跑,outbound 无法直接测量;精确测量需要 node 回报自己观测到的 gateway→node one-way。",
    )
)

ROWS.extend(throughput_rows("Node->Gateway", "Node → Gateway", "node→gw", TRANSPORT_GW_NODE))
ROWS.extend(
    latency_rows(
        "Node->Gateway",
        "OneWayLatency",
        "这里测量的是 Node → Gateway 单向 wire latency:node 把 WS frame 发出的瞬间(frame.sentAt)到 gateway 收到的瞬间之间的差值。",
        "打点在 protocol-trace-store.captureTrace:recvTs − sentAt,存到 trace.oneWayLatencyMs,UI 端按 source 聚合。前提是时钟 NTP 同步。",
    )
)

# ---------------------------------------------------------------------------
# Agent ↔ Model
# ---------------------------------------------------------------------------
ROWS.extend(throughput_rows("Agent->Model", "Agent → Model", "agent→model", TRANSPORT_AGENT_MODEL))

# Agent → Model Requests section (new)
ROWS.extend(
    [
        (
            "Agent->Model.Requests.Total",
            "自 session 开始以来 Agent 向 Model 发出的 request 次数 —— agent 每一次把 prompt 交给 LLM 推理。",
            "controller 侧持久累加器:每条 stream=lifecycle 且 phase ∈ {start, request} 的事件 +1。按 trace id 去重,不受 ring buffer eviction 影响。",
            "一次 agent task 通常几次到几十次;长探索任务可过百。",
            "每次 request 的 payload 取自事件的 data.requestSize,写入 trace.payloadSize。Total Requests × Avg Payload ≈ agent→model forward 累计字节。",
        ),
        (
            "Agent->Model.Requests.PeakPayload",
            "所有 request 中最大的一次 payload。反映了 transcript + system prompt + tool catalog 的峰值。",
            "max(requestSize),累加器在入队时与当前值比较。",
            "初轮通常 < 500 KB;后轮 transcript 累积可上 MB;开启 cache control 时 cache-create 的那一条常是 peak。",
            "transcript 越长、tool result 越多,后续 request 的 input 越大。",
        ),
        (
            "Agent->Model.Requests.AvgPayload",
            "所有 request payload 大小的算术平均。",
            "round(totalSize / total)。",
            "短任务可能几十到几百 KB;长 agent run 平均可到几 MB。",
            "peak ≫ avg → 分布长尾明显;peak ≈ avg → request 大小均匀。",
        ),
        (
            "Agent->Model.Requests.LatestRequest",
            "最近一次 Agent → Model request 的时间与 payload 大小。",
            "在每次 ingestion 中更新 latestTs = max(ts),同步更新 latestSize;卡片显示 HH:MM:SS 与 formatBytes(latestSize)。",
            "活跃 agent 在最近几秒内;agent 在执行 tool 或空闲时可能是几分钟前。",
            "可用于和 Protocol 序列图上最后一条 agent→llm 箭头对齐验证。",
        ),
    ]
)

# TTFT latency on agent→model
ROWS.extend(
    latency_rows(
        "Agent->Model",
        "TTFT",
        "这里测量的是 TTFT(time-to-first-token):agent 把请求发出去之后,到 model 第一次回流 assistant 内容之间的等待时间。",
        "computeAgentLlmTtft:对每个 runId,以 lifecycle start 或上一次 tool end/result 为起点,以下一条 stream === 'assistant' 的 trace 为终点,差值即一次 TTFT。多次 LLM 调用分别打点为 TTFT #1、#2 …",
    )
)

# Model → Agent Responses section (new)
ROWS.extend(throughput_rows("Model->Agent", "Model → Agent", "model→agent", TRANSPORT_MODEL_AGENT))

ROWS.extend(
    [
        (
            "Model->Agent.Responses.TotalSSEEvents",
            "自 session 开始以来 Model → Agent 方向累计的 assistant stream 事件数 —— 对应 LLM provider 的 Server-Sent Events(每帧 = 一个 delta token / tool_use 增量 / stop 标记)。",
            "controller 持久累加:每条 source=llm, target=agent, stream=assistant 的 trace +1。按 id 去重,不受 ring buffer eviction 影响。",
            "单次 LLM 调用几十到几百帧;长 agent run 可达上万。",
            "SSE 频率反映 provider 流式吐字节奏:generation 稳定 vs 长间隔 / rate-limit 都能先从这里看出来。",
        ),
        (
            "Model->Agent.Responses.AvgEventsPerSec",
            "SSE 事件平均每秒多少帧。",
            "total / ((lastTs − firstTs) / 1000),仅当 ≥ 2 帧时有值。",
            "常见 provider 10–60 帧/秒;有 tool 空档的任务会偏低(分母含空档)。",
            "多次 LLM call 中间 tool 执行的空档也会被算入“平均”,因此通常低于单次 generation burst 的瞬时 rate。",
        ),
        (
            "Model->Agent.Responses.PeakPayload",
            "单个 SSE 事件的最大 payload。通常 initial message_start / 完整 tool_use block 是 peak。",
            "max(payloadSize),在 ingestion 时与当前 peak 比较。",
            "delta 帧通常 < 1 KB;tool_use 一次发完可达几十 KB。",
            "ring buffer 会把 assistant stream payload 截断到 { _slim: true, data: { text } },但 payloadSize 是在截断前估算的,反映真实 wire 大小。",
        ),
        (
            "Model->Agent.Responses.AvgPayload",
            "所有 SSE 事件 payload 大小的算术平均。",
            "round(totalSize / total)。",
            "纯 delta 流平均几百字节;混合 tool_use 的流平均 1–5 KB。",
            "Avg SSE × Avg Events/Sec ≈ Model→Agent 的平均 bytes/s,可与 Throughput.Average 交叉验证。",
        ),
    ]
)

# Generation latency on model→agent
ROWS.extend(
    latency_rows(
        "Model->Agent",
        "Generation",
        "这里测量的是 assistant 流式输出本身的持续时间:从 model 回出第一段 assistant 内容,到下一个 tool call(或 lifecycle end)之间的窗口。",
        "computeAgentLlmGeneration:从一段 assistant burst 的第一条流式事件到最后一条之间的时间(在下一个 tool call 或 lifecycle end 之前)。一次 burst = 一次 generation 样本。",
    )
)


# ---------------------------------------------------------------------------
# Write workbook
# ---------------------------------------------------------------------------

def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Protocol Monitor Metrics"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(vertical="center", horizontal="left", wrap_text=True)

    for col, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    body_align = Alignment(vertical="top", horizontal="left", wrap_text=True)
    metric_font = Font(name="Menlo", size=10)

    # Alternating row fill for readability
    zebra_fill = PatternFill("solid", fgColor="F3F4F6")

    section_prefixes = [
        "UsageOverview",
        "Operator->Gateway",
        "Gateway->Operator",
        "Gateway->Node",
        "Node->Gateway",
        "Agent->Model",
        "Model->Agent",
    ]

    current_section = None
    row_idx = 2
    for metric, meaning, how, typical, notes in ROWS:
        # Determine section (first token before '.')
        section = metric.split(".")[0]
        if section != current_section:
            current_section = section

        # Zebra based on section index to visually group
        section_index = section_prefixes.index(section) if section in section_prefixes else 0
        use_zebra = section_index % 2 == 1

        values = [metric, meaning, how, typical, notes]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.alignment = body_align
            if col == 1:
                cell.font = metric_font
            if use_zebra:
                cell.fill = zebra_fill
        row_idx += 1

    # Column widths — metric column tight, narrative columns wider
    widths = {1: 40, 2: 60, 3: 60, 4: 40, 5: 60}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Freeze header
    ws.freeze_panes = "A2"

    # Row height auto-ish by setting a reasonable default
    for r in range(2, row_idx):
        ws.row_dimensions[r].height = 80

    out_path = os.path.expanduser("~/Desktop/protocol-monitor-metrics.xlsx")
    wb.save(out_path)
    print(f"Wrote {out_path} with {row_idx - 2} metric rows.")


if __name__ == "__main__":
    main()
